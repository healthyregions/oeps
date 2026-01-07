from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

from ..handlers import Registry, Metadata

def parse_attribute_from_variable(
        attribute: str,
        variable: dict,
        metadata: Metadata,
    ) -> str:

    v = variable["info"]
    if attribute == "Longitudinal":
        if v.longitudinal:
            return "x"
    elif attribute == "Analysis":
        if v.analysis:
            return "x"
    elif attribute == "Theme":
        return metadata.theme
    elif attribute == "Construct":
        return metadata.construct2
    elif attribute in variable.get("years", []):
        return "x"
    elif attribute == "Title":
        return v.title
    elif attribute == "Variable":
        return v.name
    elif attribute == "Metadata Location":
        return metadata.url
    elif attribute == "Source":
        return metadata.source
    elif attribute == "Source Long":
        return metadata.source_long
    elif attribute.lower() in v:
        return v.__dict__.get(attribute.lower())

    return ""

def create_data_dictionary(registry: Registry, summary_level: str, destination: Path):
    """Generate MS Excel formatted data dictionaries for all content in a given summary level."""

    destination.mkdir(exist_ok=True)
    out_path = Path(destination, f"{summary_level[0].upper()}_Dict.xlsx")

    tabular = [
        i
        for i in registry.table_sources.values()
        if registry.geodata_sources[i.geodata_source].summary_level.name == summary_level
    ]

    # get all variables (fields) that are in any of these tables
    fields = []
    for table in tabular:
        for i in registry.variables.values():
            if table.name in i.table_sources:
                fields.append(i)

    ordered = []
    for theme in registry.theme_tree.keys():
        matched = []
        for field in fields:
            if registry.metadata[field.metadata].theme == theme:
                matched.append(field)
        ordered += sorted(matched, key=lambda i: (registry.metadata[i.metadata].construct2, i.name))

    all_variables = {}
    for f in ordered:
        years = set([registry.table_sources[i].data_year for i in f.table_sources])
        all_variables[f.name] = {"years": years, "info": f}

    years_list = set()
    for v in all_variables.values():
        for y in v["years"]:
            years_list.add(y)

    headers = {"Theme": 15, "Construct": 20}
    for y in sorted(years_list):
        headers[y] = 5
    headers.update(
        {
            "Analysis": 10,
            "Longitudinal": 10,
            "Variable": 20,
            "Title": 30,
            "Description": 25,
            "Metadata Location": 25,
            "Source": 25,
            "Source Long": 25,
            "Type": 25,
            "Example": 25,
            "Comments": 100,
        }
    )

    wb = Workbook()
    ws = wb.active
    ws.append(list(headers.keys()))

    ft = Font(bold=True, name="Calibri")
    for row in ws["A1:Z1"]:
        for cell in row:
            cell.font = ft

    for v in all_variables.values():
        metadata = registry.metadata[v['info'].metadata]
        row = [parse_attribute_from_variable(i, v, metadata) for i in headers.keys()]
        ws.append(row)

    for n, k in enumerate(headers.keys()):
        ws.column_dimensions[get_column_letter(n + 1)].width = headers[k]

    # Save the file
    wb.save(out_path)
