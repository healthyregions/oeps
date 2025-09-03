import os
import csv
from enum import Enum
from pathlib import Path
import shutil
from warnings import warn

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import pandas as pd
import geopandas as gpd

from oeps.config import REGISTRY_DIR, TEMP_DIR, DATA_DIR
from oeps.utils import load_json, write_json

THEME_ORDER = [
    "Geography",
    "Social",
    "Environment",
    "Economic",
    "Policy",
    "Outcome",
    "Composite",
]

summary_lookup = {
    "state": {
        "code": "040",
        "geoid_length": 2,
    },
    "county": {
        "code": "050",
        "geoid_length": 5,
    },
    "zcta": {
        "code": "860",
        "geoid_length": 5,
    },
    "tract": {
        "code": "140",
        "geoid_length": 11,
    },
}


class SummaryLevel(Enum):
    state = "state"
    county = "county"
    zcta = "zcta"
    tract = "tract"


class TableSource:
    def __init__(
        self,
        name: str,
        with_data: bool = False,
        registry_dir: Path = REGISTRY_DIR,
        registry: "Registry" = None,
    ):
        if not registry:
            registry = Registry(registry_dir, quiet=True)
        self.registry = registry
        if name not in self.registry.table_sources:
            raise Exception(f"invalid table source name: {name}")
        self.name = name
        self.schema = self.registry.table_sources[name]
        self.summary_level = self.registry.geodata_sources[
            self.schema["geodata_source"]
        ]["summary_level"]

        self.path = Path(DATA_DIR, "tables", f"{self.name}.csv")

        temp_out = Path(TEMP_DIR, "tables")
        temp_out.mkdir(parents=True, exist_ok=True)
        self.temp_path = Path(temp_out, f"{self.name}.csv")

        self.merge_columns = []

        if with_data:
            self.df = self.load_dataframe()

    def load_dataframe(self) -> pd.DataFrame:
        """Load this TableSource's CSV data into a pandas DataFrame"""

        return pd.read_csv(self.get_path())

    def get_path(self) -> str | Path:
        path = self.schema["path"]
        if not path.startswith("http"):
            path = Path(DATA_DIR, path)
        return path

    def set_data_types(self):
        """set integer columns properly based on registry def of variables.
        Raise a warning if this process fails on a given column."""

        for col in self.df.columns:
            if self.registry.variables[col]["type"] == "integer":
                try:
                    self.df[col] = self.df[col].astype("Int64")
                except pd.errors.IntCastingNaNError:
                    warn(
                        message=f"Failed to coerce column {col} to Integer due to presence of NA or inf values in the dataset. Continuing without coercing."
                    )

    def stage_incoming_csv(self, path: Path) -> pd.DataFrame:
        """Load an incoming CSV to pandas dataframe, and create HEROP_ID
        along the way if possible."""

        df = pd.read_csv(path)

        lvl = summary_lookup[self.summary_level]

        ## all good if HEROP_ID already exists in the data frame
        test_unique_id = "HEROP_ID"
        if "HEROP_IP" not in df.columns:
            for id in ["GEOID", "GEO ID", "GEO_ID", "FIPS", "ZCTA5"]:
                if id in df.columns:
                    df["HEROP_ID"] = f"{lvl['code']}US" + df[id].astype(str).str.zfill(
                        lvl["geoid_length"]
                    )
                    test_unique_id = id

        ## make sure whatever column that is used for the join ID is unique
        if not pd.Series(df[test_unique_id]).is_unique:
            raise Exception(
                f"There are duplicate {test_unique_id} values in the input CSV. {test_unique_id} must be unique across all rows."
            )

        print("incoming data frame loaded")
        print(df)

        if "HEROP_ID" in df.columns:
            self.staged_df = df
        ## if it wasn't added above based on other incoming fields, abort process
        else:
            raise Exception(
                "input data frame must have one of these fields: HEROP_ID, GEOID, GEO ID, GEO_ID, FIPS, ZCTA5"
            )

    def validate_incoming_csv(self, verbose: bool = False):
        """Compares the columns in the incoming CSV and against variables
        in the registry and columns in the existing data for this TableSource."""

        ## determine which columns to take from the incoming dataframe
        matched = [i for i in self.staged_df.columns if i in self.registry.variables]
        missed = [i for i in self.staged_df.columns if i not in self.registry.variables]
        overlap = [i for i in matched if i in self.df.columns if not i == "HEROP_ID"]
        new = [i for i in matched if i not in self.df.columns and i not in overlap]

        print(f"{len(matched)} columns match to variables already in the registry")
        print(
            f"{len(missed)} columns are not yet in the registry and will be ignored. List of unmatched columns:"
        )
        for i in missed:
            print(f"  {i}")
        print(
            f"{len(overlap)} columns in the incoming dataset already exist in the target"
        )
        if overlap:
            print("  -- overlapping columns from incoming data will be ignored.")

        print(f"{len(new)} new column(s) will be added to the existing CSV")

        self.merge_columns = new

    def merge_incoming_csv(self, dry_run: bool = False):
        source_df = self.staged_df

        source_df_trim = source_df[["HEROP_ID"] + self.merge_columns]
        print(source_df_trim)
        print(f"shape of incoming data: {source_df_trim.shape}")
        self.df = pd.merge(self.df, source_df_trim, how="inner", on="HEROP_ID")

        self.set_data_types()

        self.df.to_csv(self.temp_path, index=False)

        if not dry_run:
            self.df.to_csv(self.path, index=False)
            self.registry.sync_variable_table_sources(table_source=self)

    def get_variable_data(self, name):
        return pd.DataFrame(self.df, columns=["HEROP_ID", name])

    def delete_variable_data(self, name):
        self.df = self.df.drop(name, axis=1)
        self.df.to_csv(self.path, index=False)

    def merge_df(self, incoming_df: pd.DataFrame, overwrite: bool = False):
        new_vars = [i for i in incoming_df.columns if not i == "HEROP_ID"]
        dupe_vars = [i for i in new_vars if i in self.df.columns]
        if len(dupe_vars) > 0:
            if not overwrite:
                raise Exception(
                    "Incoming variables exist in the dataframe, cancelling merge. Use overwrite=True to continue."
                )
            for var in dupe_vars:
                self.delete_variable_data(var)

        self.df = pd.merge(self.df, incoming_df, how="inner", on="HEROP_ID")

        self.set_data_types()

        self.df.to_csv(self.path, index=False)


class Registry:
    def __init__(self, directory: Path = REGISTRY_DIR, quiet: bool = False):
        self.directory = directory

        ## load in this order so that some attributes can be cascaded
        self.geodata_sources = self._load_geodata_sources()
        self.table_sources = self._load_table_sources()
        self.metadata = self._load_metadata()
        self.theme_tree = self._load_theme_tree()
        self.variables = self._load_variables()

        if not quiet:
            print(
                f"registry initialized | variables: {len(self.variables)}, tables: {len(self.table_sources)}, geodata: {len(self.geodata_sources)}"
            )

    def _load_geodata_sources(self) -> dict:
        """Creates a lookup of all geodata sources in the registry."""

        output = {}
        paths = Path(self.directory, "geodata_sources").glob("*.json")
        for path in paths:
            data = load_json(path)
            output[data["name"]] = data
        return output

    def _load_table_sources(self) -> dict:
        """Creates a lookup of all table sources in the registry."""

        output = {}
        paths = Path(self.directory, "table_sources").glob("*.json")
        for path in paths:
            resource = load_json(path)
            resource_id = resource["name"]

            schema = {
                "primaryKey": "HEROP_ID",
                "missingValues": ["NA"],
                "foreignKeys": [],
                "fields": [],
            }

            ## if there is a geodata_source (which there should be), then
            ## attach information from it to this table_source
            join_resource = resource.get("geodata_source")
            if join_resource:
                schema["foreignKeys"].append(
                    {
                        "fields": "HEROP_ID",
                        "reference": {
                            "resource": join_resource,
                            "fields": "HEROP_ID",
                        },
                    }
                )
                geodata = self.geodata_sources[join_resource]
                resource["summary_level"] = geodata["summary_level"]

            resource["schema"] = schema
            output[resource_id] = resource

        return output

    def _load_variables(self) -> dict:
        """Creates a lookup of all variables."""

        variables = {}
        for path in Path(self.directory, "variables").glob("*.json"):
            variables[path.stem] = load_json(path)

        ## iterate all table_sources and add field lists to their schema
        ## using these variables.
        for k, v in self.table_sources.items():
            v["schema"]["fields"] = [
                i for i in variables.values() if k in i["table_sources"]
            ]

        ## add year
        for v in variables.values():
            sources = [
                i
                for i in self.table_sources.values()
                if i["name"] in v.get("table_sources", [])
            ]
            v["years"] = list(set([i["year"] for i in sources]))

        for i in variables.values():
            if i["metadata"] not in self.metadata:
                print(
                    f"WARNING: variable {i['name']} references unknown metadata {i['metadata']}"
                )

        return variables

    def _load_metadata(self):
        metadata = {}
        for path in Path(self.directory, "metadata").glob("*.json"):
            metadata[path.stem] = load_json(path)

        return metadata

    def _load_theme_tree(self):
        ## construct theme_tree based on sort order and then content in metadata
        tree = {k: {} for k in THEME_ORDER}
        for k, md in self.metadata.items():
            valid = True
            for key in ["id", "theme", "construct", "proxy", "url"]:
                if key not in md or not md[key]:
                    print(f"WARNING: metadata entry {k} is missing value: '{key}'")
                    valid = False
            if not valid:
                continue

            n = md["id"]
            t = md["theme"]
            c = md["construct"]

            ## theme must be one of the preset options
            if md["theme"] not in tree:
                print(f"WARNING: metadata entry {n} has invalid theme {t}")
            else:
                if c not in tree[t]:
                    tree[t][c] = [n]
                else:
                    tree[t][c].append(n)

        return tree

    def sync_variable_table_sources(self, table_source: TableSource):
        for var in self.variables.values():
            ## first remove this table source from all variables in the registry
            var["table_sources"] = [
                i for i in var["table_sources"] if not i == table_source.name
            ]
            ## now add this table source if the variable is in df.columns
            if var["name"] in table_source.df.columns:
                var["table_sources"].append(table_source.name)
            ## sort all table_sources
            var["table_sources"].sort()
        self.write_variables()

    def remove_variable(self, variable_name: str):
        os.remove(Path(REGISTRY_DIR, "variables", f"{variable_name}.json"))
        self._load_variables()

    def write_variables(self):
        for v in self.variables.values():
            if "years" in v:
                del v["years"]
            write_json(v, Path(self.directory, "variables", f"{v['name']}.json"))

    def get_all_sources(self):
        sources = list(self.table_sources.values())
        sources += list(self.geodata_sources.values())
        return sources

    def find_table_source(self, variable_name, summary_level, year_up_to: int = None):
        """Given the input variable name and summary level, return the latest
        matching table_source.

        year_up_to: Optionally provide a year cutoff after which table_sources
        will be ignored."""

        variable = self.variables[variable_name]
        matching_sources = []
        for ts in variable["table_sources"]:
            ts_full = self.table_sources.get(ts)
            if ts_full and ts_full["summary_level"] == summary_level:
                matching_sources.append(ts_full)

        if len(matching_sources) == 0:
            return None

        matching_sources.sort(key=lambda d: d["year"])
        if year_up_to is None:
            return matching_sources[-1]
        else:
            use_source = matching_sources[0]
            for source in matching_sources:
                if int(source["year"]) > year_up_to:
                    break
                use_source = source
            return use_source

    def create_data_dictionaries(self, destination: Path = None):
        """Generate MS Excel formatted data dictionaries for all content."""

        if not destination:
            destination = Path(DATA_DIR, "dictionaries")
        destination.mkdir(exist_ok=True)

        summary_levels = {}
        for geodata in self.geodata_sources.values():
            summary_levels[geodata["summary_level"][0].upper()] = geodata[
                "summary_level"
            ]

        for id, label in summary_levels.items():
            tabular = [
                i
                for i in self.table_sources.values()
                if self.geodata_sources[i["geodata_source"]]["summary_level"] == label
            ]

            # get all variables (fields) that are in any of these tables
            fields = []
            for table in tabular:
                for i in self.variables.values():
                    if table["name"] in i.get("table_sources", {}):
                        fields.append(i)

            ordered = []
            for theme in self.theme_tree.keys():
                matched = []
                for field in fields:
                    metadata_id = field.get("metadata", "")
                    metadata_id = metadata_id
                    if metadata_id != "":
                        if self.metadata[metadata_id]["theme"] == theme:
                            matched.append(field)
                ordered += sorted(matched, key=lambda i: i["metadata"])

            all_variables = {}
            for f in ordered:
                all_variables[f["name"]] = {"years": f["years"], "info": f}

            years_list = set()
            for v in all_variables.values():
                for y in v["years"]:
                    years_list.add(y)

            headers = {"Theme": 15}
            for y in sorted(years_list):
                headers[y] = 5
            headers.update(
                {
                    "Analysis": 10,
                    "Longitudinal": 10,
                    "Variable": 20,
                    "Title": 30,
                    "Description": 25,
                    "Metadata Location": 25,
                    "Source": 25,
                    "Source Long": 25,
                    "OEPS v1 Table": 25,
                    "Type": 25,
                    "Example": 25,
                    "Data Limitations": 25,
                    "Comments": 100,
                }
            )

            wb = Workbook()
            ws = wb.active
            ws.append(list(headers.keys()))

            ft = Font(bold=True, name="Calibri")
            for row in ws["A1:Z1"]:
                for cell in row:
                    cell.font = ft

            def parse_attribute_from_variable(attribute, variable):
                v = variable["info"]
                if attribute == "Longitudinal":
                    if v["longitudinal"]:
                        return "x"
                elif attribute == "Analysis":
                    if v["analysis"]:
                        return "x"
                elif attribute == "Theme":
                    return self.metadata[v["metadata"]]["theme"]
                elif attribute in variable.get("years", []):
                    return "x"
                elif attribute == "Title":
                    return v.get("title")
                elif attribute == "Variable":
                    return v.get("name")
                elif attribute == "Metadata Location":
                    return self.metadata[v["metadata"]]["url"]
                elif attribute == "Data Limitations":
                    return v.get("constraints")
                elif attribute == "Source Long":
                    return v.get("source_long")
                elif attribute == "OEPS v1 Table":
                    return v.get("oeps_v1_table")
                elif attribute.lower() in v:
                    return v.get(attribute.lower())

                return ""

            for v in all_variables.values():
                row = [parse_attribute_from_variable(i, v) for i in headers.keys()]
                ws.append(row)

            for n, k in enumerate(headers.keys()):
                ws.column_dimensions[get_column_letter(n + 1)].width = headers[k]

            # Save the file
            wb.save(Path(destination, f"{id}_Dict.xlsx"))

    def validate(self):
        print("\n## Make dataframe lookup of all table sources")
        ts_lookup = {}
        for k, v in self.variables.items():
            for ts in v["table_sources"]:
                if ts not in self.table_sources:
                    print(f"{k} references a table_source not in the registry: {ts}")
                    variables_valid = False
                    continue

                if ts not in ts_lookup:
                    ts_lookup[ts] = TableSource(ts, with_data=True, registry=self)

        print("\n## Check integrity of all CSV files")
        for id, ts in ts_lookup.items():
            with open(ts.get_path(), "r") as o:
                reader = csv.reader(o)
                headers = next(reader)
            for i in headers:
                if headers.count(i) > 1:
                    print(f"duplicate column {i} in CSV for {id}")
                if id not in self.variables[i]["table_sources"]:
                    print(
                        f"column {i} present in CSV for {id} but not listed as table-source"
                    )

        print("\n## Checking variables against table source data")

        variables_valid = True
        for k, v in self.variables.items():
            for id in v["table_sources"]:
                ts = ts_lookup.get(id)
                if ts is None:
                    continue

                if k not in ts.df.columns:
                    print(
                        f"{k} should be in table_source {ts} but that column is not present in the CSV"
                    )
                    variables_valid = False
        if variables_valid:
            print("all good")

        print("\n## Checking geodata sources")
        geodata_valid = True
        for k, v in self.geodata_sources.items():
            try:
                SummaryLevel(v["summary_level"])
            except Exception:
                print(k)
                geodata_valid = False
        if geodata_valid:
            print("all good")

    def create_table_source(
        self, name: str, geodata_source: str, dry_run: bool = False
    ) -> TableSource:
        summary_level, year = name.split("-")
        geog_summary_level = geodata_source.split("-")[0]

        ## validate these components
        SummaryLevel(summary_level)
        int(year)

        gs = self.geodata_sources[geodata_source]

        file_name = f"{name}.csv"
        temp_path = Path(TEMP_DIR, "tables", file_name)

        schema = {
            "bq_dataset_name": "tabular",
            "bq_table_name": name,
            "name": name,
            "path": f"tables/{file_name}",
            "format": "csv",
            "mediatype": "text/csv",
            "title": name,
            "description": f"This CSV aggregates all OEPS data values from {year} at the {gs['summary_level']} level.",
            "year": str(year),
            "geodata_source": geodata_source,
        }

        ## write the definition to a new JSON file
        if not dry_run:
            outpath = Path(self.directory, "table_sources", f"{name}.json")
            write_json(schema, outpath)

        ## now create a dummy CSV with all key variables, based on the geodata_source
        temp_path = Path(TEMP_DIR, "tables", file_name)
        gdf = gpd.read_file(gs["path"])[["HEROP_ID"]]

        foreign_key = "ZCTA5" if geog_summary_level == "zctas" else "FIPS"
        gdf[foreign_key] = gdf.HEROP_ID.apply(lambda x: x[5:])

        print("new data table:")
        print(gdf)

        gdf.to_csv(temp_path, index=False)

        if not dry_run:
            local_path = Path(DATA_DIR, schema["path"])
            shutil.copy(temp_path, local_path)
