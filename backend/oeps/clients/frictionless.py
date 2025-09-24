import os
import csv
import shutil
import pandas as pd
from pathlib import Path

from frictionless import validate
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

from ..clients.s3 import upload_to_s3
from ..registry.handlers import Registry
from ..registry.models import GeodataSourceModel, GEOGRAPHY_LOOKUP
from ..utils import load_json, write_json, download_file


class DataPackage:
    def __init__(self, path: Path = None):
        self.path = path
        self.registry = None
        self.rules_rows = []
        self.verbose = False
        self.schema = {
            "profile": "data-package",
            "name": None,
            "title": None,
            "homepage": "https://oeps.healthyregions.org",
            "licenses": [
                {
                    "name": "ODC-PDDL-1.0",
                    "path": "http://opendatacommons.org/licenses/pddl/",
                    "title": "Open Data Commons Public Domain Dedication and License v1.0",
                }
            ],
            "resources": [],
        }

    def check_geodata_sources_in_rules_file(self):
        gs = dict()
        for row in self.rules_rows:
            ts = self.registry.table_sources[row["table_source"]]
            gs[ts.geodata_source] = gs.get(ts.geodata_source, []) + [row["name"]]

        return gs

    def add_geodata_source_to_package(self, gs: GeodataSourceModel):

        gs_resource = gs.to_frictionless_resource()

        gs_url = gs_resource.pop("path")
        gs_resource["path"] = f"data/{Path(gs_url).name}"
        download_file(gs_url, Path(self.path, gs_resource["path"]))
        gs_schema = gs_resource.pop('schema')
        gs_resource["schema"] = f"schemas/{gs.name}.json"
        write_json(gs_schema, Path(self.path, gs_resource["schema"]))
        self.schema["resources"].append(gs_resource)

    def create_from_rules(
        self,
        registry: Registry,
        rules_dir: Path,
        check_rules: bool = False,
        zip_output: bool = False,
        upload: bool = False,
        no_cache: bool = False,
        skip_foreign_keys: bool = False,
        run_validation: bool = True,
        verbose: bool = False,
    ):
        """Single command to generate an output data package, based on a set
        of CSVs that hold inclusion rules for each variable."""

        self.registry = registry
        self.verbose = verbose
        self.path.mkdir(exist_ok=True, parents=True)
        s_path = Path(self.path, "schemas")
        s_path.mkdir(exist_ok=True)
        d_path = Path(self.path, "data")
        d_path.mkdir(exist_ok=True)

        rules_name = rules_dir.name

        self.schema["name"] = f"oeps-{rules_name.lower()}"
        self.schema["title"] = f"Opioid Environment Policy Scan (OEPS) - {rules_name}"

        geodata_sources = []
        for rules_file in rules_dir.glob("*.csv"):
            if rules_file.stem not in GEOGRAPHY_LOOKUP:
                continue
            print(f"processing {rules_file.name}")
            with open(rules_file, "r") as o:
                reader = csv.DictReader(o)
                ## filter out variables that don't have a "table_source" value
                self.rules_rows = [i for i in reader if i["table_source"]]

            gs_lookup = self.check_geodata_sources_in_rules_file()
            if len(gs_lookup.keys()) > 1:
                print("not all variable years in this file use the same geodata source.")
                print("this must be corrected before the file can be processed.")

                fewer_vars = None
                fewer_ct = 10000
                for k, v in gs_lookup.items():
                    if len(v) < fewer_ct:
                        fewer_ct = len(v)
                        fewer_vars = v
                    print(k, len(v))
                print("outliers:\n" + '\n'.join(fewer_vars))
                continue

            if check_rules:
                continue

            gs = self.registry.geodata_sources.get(list(gs_lookup.keys())[0])
            geodata_sources.append(gs)

            ts_resource = {
                "name": rules_file.stem,
                "title": rules_file.stem,
                "format": "csv",
                "mediatype": "text/csv",
                "path": f"data/{rules_file.name}",
                "schema": f"schemas/{rules_file.stem}.json",
            }
            resource_schema = {
                "fields": [
                    {
                        "title": "HEROP_ID",
                        "name": "HEROP_ID",
                        "type": "string",
                        "example": "050US01001",
                        "description": "A derived unique id corresponding to the relevant geographic unit.",
                        "metadata": "Geographic_Boundaries"
                    }
                ]
            }
            if not skip_foreign_keys:
                resource_schema["foreignKeys"] = [{
                    "fields": "HEROP_ID",
                    "reference": {
                        "resource": gs.name,
                        "fields": "HEROP_ID"
                    }
                }]
            if rules_file.stem == "zcta":
                resource_schema["fields"].append({
                    "title": "ZCTA5",
                    "name": "ZCTA5",
                    "type": "string",
                    "example": "22001",
                    "description": "Zip Code for this geographic unit.",
                    "metadata": "Geographic_Boundaries"
                })
            else:
                resource_schema["fields"].append({
                    "title": "FIPS",
                    "name": "FIPS",
                    "type": "string",
                    "example": "22001",
                    "description": "FIPS code for this geographic unit.",
                    "metadata": "Geographic_Boundaries"
                })

            df = gs.get_blank_dataframe()
            for row in self.rules_rows:
                ## skip HEROP_ID and FIPS, as they are already in the blank df
                if row["name"] in ["HEROP_ID", "FIPS"]:
                    continue
                variable = self.registry.variables.get(row["name"])
                if not variable:
                    raise Exception(f"skipping variable not in registry: {row['name']}")
                ts = self.registry.table_sources[row["table_source"]]

                var_df = ts.get_variable_data([variable.name])
                df = pd.merge(df, var_df, how="left", on="HEROP_ID")

                field_def = variable.to_frictionless_field()
                field_def["data_year"] = ts.data_year
                resource_schema["fields"].append(field_def)

            print(df)
            outpath = Path(d_path, rules_file.name)
            df.to_csv(outpath, index=False)

            self.schema["resources"].append(ts_resource)
            write_json(resource_schema, Path(self.path, ts_resource["schema"]))

            self.clean_data_resource(ts_resource)

        write_json(self.schema, Path(self.path, "data-package.json"))

        self.collect_metadata()

        for geodata_source in geodata_sources:
            self.add_geodata_source_to_package(geodata_source)

        self.create_data_dictionaries()

        if run_validation:
            self.validate()
        else:
            print("skipping data package validation...")

        if zip_output or upload:
            print("zipping output...")
            shutil.make_archive(
                f"{Path(self.path.parent, self.path.name)}", "zip", self.path
            )

            zip_path = self.path.with_suffix(".zip")

            if upload:
                print("uploading zip to S3...")
                upload_to_s3(zip_path, prefix="oeps", progress_bar=True)

            if not zip_output:
                print("deleting local copy of zipped output...")
                os.remove(zip_path)

        return self.path

    def validate(self):

        print("\nvalidating output data package...")
        report = validate(Path(self.path, "data-package.json"), skip_errors=["type-error"])

        print("VALIDATION REPORT SUMMARY:")
        for t in report.tasks:
            print(t.name, t.stats["errors"])
            if self.verbose:
                if len(t.errors) > 0:
                    for e in t.errors:
                        print(e)

        print(f"Totals: {report.stats}")

        report.to_json(Path(self.path, "error-report.json"))

    def collect_metadata(self):

        urls = set()
        for schema_file in Path(self.path, "schemas").glob("*.json"):
            schema = load_json(schema_file)
            for field in schema["fields"]:
                md_id = field.get("metadata")
                if md_id:
                    urls.add(self.registry.metadata[md_id].url)

        if len(urls) > 0:
            md_dir = Path(self.path, "metadata")
            md_dir.mkdir(exist_ok=True)

        download_url_base = "https://raw.githubusercontent.com/healthyregions/oeps/refs/heads/main/metadata/"
        for url in urls:
            file_name = Path(url).name
            raw_url = f"{download_url_base}{file_name}"
            local_path = Path(md_dir, file_name)
            download_file(raw_url, local_path)

    def create_data_dictionaries(self):

        res_names = []
        var_rows = {}

        print("creating data dictionaries...")
        for resource in self.schema['resources']:
            if resource['format'] != "csv":
                continue

            res_names.append(resource["name"])
            schema_path = Path(self.path, resource["schema"])
            s = load_json(schema_path)
            for field in s["fields"]:
                if field["name"] in ["HEROP_ID", "FIPS", "ZCTA5"]:
                    continue

                md_id = field.get("metadata")
                theme, con, url = "", "", ""
                if md_id:
                    theme = self.registry.metadata[md_id].theme
                    con = self.registry.metadata[md_id].construct2
                    url = self.registry.metadata[md_id].url
                if field["name"] not in var_rows:
                    var_rows[field["name"]] = {
                        "Name": field["name"],
                        "Title": field["title"],
                        "Theme": theme,
                        "Construct": con,
                        "Metadata": url,
                    }
                var_rows[field["name"]][resource["name"]] = field["data_year"]

        headers = {
            "Name": 15,
            "Title": 30,
        }
        for i in res_names:
            headers[i] = 15
        headers.update({
            "Theme": 15,
            "Construct": 20,
            "Metadata": 45,
        })

        wb = Workbook()
        ws = wb.active

        ws.append(list(headers.keys()))

        for variable in sorted(var_rows.values(), key=lambda x: (x["Theme"], x["Construct"], x["Name"])):
            row = []
            for col in headers.keys():
                row.append(variable.get(col))
            ws.append(row)

        ft = Font(bold=True, name="Calibri")
        for row in ws["A1:Z1"]:
            for cell in row:
                cell.font = ft
        for n, k in enumerate(headers.keys()):
            ws.column_dimensions[get_column_letter(n + 1)].width = headers[k]

        out_path = Path(self.path, "data-dictionary.xlsx")
        wb.save(out_path)

    def clean_data_resource(self, res):
        print(f"cleaning data for {res['name']}")

        data_path = self.path / res["path"]
        schema_path = self.path / res["schema"]

        schema = load_json(schema_path)
        fields = {i["name"]: i for i in schema["fields"]}
        df = pd.read_csv(data_path)

        # create new dataframe with only fields as defined in the schema
        # (this takes care of ordering the fields properly as well)
        clean_df = df[fields.keys()]

        # set all dtypes to generic "object" so they can hold "" for NaN
        clean_df = clean_df.astype(object)

        # convert all NaN to "", use this context and infer_objects to avoid a warning
        # see: https://stackoverflow.com/a/78066237/3873885
        with pd.option_context("future.no_silent_downcasting", True):
            clean_df = clean_df.fillna("").infer_objects(copy=False)

        # iterate all fields and if integer, cast to int to remove .0
        for k, v in fields.items():
            if v["type"] == "integer":
                clean_df[k] = clean_df[k].apply(lambda x: int(x) if x else "")
            if v["type"] == "boolean":
                clean_df[k] = clean_df[k].apply(lambda x: True if x else False)

        clean_df.to_csv(data_path, index=False)
